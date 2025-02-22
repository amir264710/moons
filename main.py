import tkinter as tk
from tkinter import filedialog, messagebox
from PIL import Image, ImageTk, Image as PILImage
import csv
import os
from fpdf import FPDF
import openpyxl
from openpyxl.drawing.image import Image as XLImage
import shutil
import random
import string

def generate_random_string(length=6):
    """Generate a random string of given length."""
    characters = string.ascii_letters + string.digits  # Letters (uppercase & lowercase) + digits
    return ''.join(random.choices(characters, k=length))

# Use the appropriate resampling method based on your Pillow version
try:
    resampling = Image.Resampling.LANCZOS
except AttributeError:
    resampling = Image.ANTIALIAS

APP_FOLDER = r"C:\WareHouse_Stock"
CSV_FILE = f"{APP_FOLDER}/data.csv"

# Global variables to hold our records and state
records = []            # List of record dictionaries
current_record_index = None  # Index of the record currently being viewed

# Variables for input mode image
input_image_path = None
input_image_photo = None

# Variable to hold the currently displayed image in view mode (to keep reference)
view_image_photo = None

# Define a default font for Tkinter that supports Persian (e.g. Tahoma)
default_font = ("Tahoma", 10)

def copy_file(source_path, destination_folder, new_filename):
    """
    Copies a file from source_path to destination_folder with a new name but keeps the same extension.

    :param source_path: Path of the file to be copied.
    :param destination_folder: Folder where the file should be pasted.
    :param new_name_without_extension: The new name for the copied file (without extension).
    """
    try:
        # Ensure the destination folder exists
        os.makedirs(destination_folder, exist_ok=True)

        # Get the original file extension
        file_extension = os.path.splitext(source_path)[1]

        # Construct the new file path with the same extension
        new_filename = new_filename + file_extension
        destination_path = os.path.join(destination_folder, new_filename)

        # Copy the file
        shutil.copy2(source_path, destination_path)
        return destination_path
    except Exception as e:
        print(f"Error copying file: {e}")

def load_records():
    global records, current_record_index
    records = []
    if os.path.isfile(CSV_FILE):
        with open(CSV_FILE, newline='', encoding='utf-8') as csvfile:
            reader = csv.DictReader(csvfile)
            for row in reader:
                records.append(row)
    if records:
        # Start with the latest record (last in the file)
        current_record_index = len(records) - 1
    else:
        current_record_index = None

def save_all_records():
    global records
    with open(CSV_FILE, "w", newline="", encoding='utf-8') as csvfile:
        fieldnames = ["Image Path", "Field1", "Field2", "Field3"]
        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
        writer.writeheader()
        for rec in records:
            writer.writerow(rec)

def show_record(index):
    global view_image_photo, current_record_index
    if index is None or not records:
        view_image_label.config(image='', text="No record available.")
        view_field1_label.config(text="")
        view_field2_label.config(text="")
        view_field3_label.config(text="")
        current_record_index = None
        return

    current_record_index = index
    rec = records[index]
    img_path = rec["Image Path"]

    if os.path.isfile(img_path):
        try:
            img = PILImage.open(img_path)
            # Resize image using thumbnail to a max size of (580,500) while preserving aspect ratio
            max_size = (580, 500)
            img.thumbnail(max_size, resampling)
            view_image_photo = ImageTk.PhotoImage(img)
            view_image_label.config(image=view_image_photo, text="")  # Clear any text
        except Exception as e:
            print("Error loading image in view mode:", e)
            view_image_label.config(image='', text=f"Error loading image: {e}")
    else:
        view_image_label.config(image='', text="Image not found")

    view_field1_label.config(text="Field 1: " + rec["Field1"])
    view_field2_label.config(text="Field 2: " + rec["Field2"])
    view_field3_label.config(text="Field 3: " + rec["Field3"])

def previous_record():
    global current_record_index
    if current_record_index is None or current_record_index <= 0:
        messagebox.showinfo("Info", "No previous record.")
    else:
        show_record(current_record_index - 1)

def next_record():
    global current_record_index
    if current_record_index is None or current_record_index >= len(records) - 1:
        messagebox.showinfo("Info", "No next record.")
    else:
        show_record(current_record_index + 1)

def delete_record():
    global current_record_index, records
    if current_record_index is None:
        messagebox.showerror("Error", "No record to delete.")
        return
    answer = messagebox.askyesno("Delete", "Are you sure you want to delete this record?")
    if answer:
        del records[current_record_index]
        save_all_records()
        if records:
            new_index = current_record_index - 1 if current_record_index > 0 else 0
            show_record(new_index)
        else:
            switch_to_input_mode()

def switch_to_input_mode():
    input_frame.tkraise()
    clear_input_form()

def switch_to_view_mode():
    view_frame.tkraise()
    if records:
        show_record(current_record_index)
    else:
        switch_to_input_mode()

def clear_input_form():
    global input_image_path, input_image_photo
    input_image_path = None
    input_image_photo = None
    input_image_label.config(image='', text="No image selected")
    entry1.delete(0, tk.END)
    entry2.delete(0, tk.END)
    entry3.delete(0, tk.END)

def select_input_image():
    global input_image_path, input_image_photo
    file_path = filedialog.askopenfilename(
        title="Select an image",
        filetypes=[("Image Files", "*.png;*.jpg;*.jpeg;*.gif")]
    )
    temp_name = generate_random_string()
    pic_path = copy_file(file_path, f"{APP_FOLDER}/pics", temp_name)
    if file_path:
        input_image_path = pic_path
        try:
            img = PILImage.open(pic_path)
            # Resize image for display in input mode using thumbnail
            max_size = (580, 500)
            img.thumbnail(max_size, resampling)
            input_image_photo = ImageTk.PhotoImage(img)
            input_image_label.config(image=input_image_photo, text="")  # Clear text
        except Exception as e:
            print("Error loading image in input mode:", e)
            input_image_label.config(image='', text=f"Error loading image: {e}")

def save_new_record():
    global records, current_record_index, input_image_path
    field1 = entry1.get()
    field2 = entry2.get()
    field3 = entry3.get()
    if not input_image_path:
        messagebox.showerror("Error", "Please select an image.")
        return
    new_record = {
        "Image Path": input_image_path,
        "Field1": field1,
        "Field2": field2,
        "Field3": field3
    }
    records.append(new_record)
    save_all_records()
    current_record_index = len(records) - 1
    messagebox.showinfo("Saved", "New record saved!")
    switch_to_view_mode()

def cancel_input():
    switch_to_view_mode()

def export_to_pdf():
    if not records:
        messagebox.showerror("Export Error", "No records available to export.")
        return
    pdf = FPDF()

    for rec in records:
        pdf.add_page()
        img_path = rec["Image Path"]
        if os.path.isfile(img_path):
            try:
                # Determine image dimensions for scaling
                pil_img = PILImage.open(img_path)
                orig_width, orig_height = pil_img.size
                pdf_img_width = 100  # Fixed width for PDF image
                scale = pdf_img_width / orig_width
                pdf_img_height = orig_height * scale
                pdf.image(img_path, x=10, y=10, w=pdf_img_width)
                pdf.set_y(10 + pdf_img_height + 10)
            except Exception as e:
                pdf.cell(0, 10, f"Error loading image: {e}", ln=1)
        else:
            pdf.cell(0, 10, "Image not found", ln=1)
            pdf.ln(20)
        pdf.cell(0, 10, "Field 1: " + rec["Field1"], ln=1)
        pdf.cell(0, 10, "Field 2: " + rec["Field2"], ln=1)
        pdf.cell(0, 10, "Field 3: " + rec["Field3"], ln=1)
    pdf_file = "records.pdf"
    try:
        pdf.output(pdf_file)
        messagebox.showinfo("Export", f"PDF exported as {pdf_file}")
    except Exception as e:
        messagebox.showerror("Export Error", f"Failed to export PDF: {e}")

def export_to_excel():
    if not records:
        messagebox.showerror("Export Error", "No records available to export.")
        return
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Records"
    ws.append(["Image", "Field1", "Field2", "Field3"])
    row = 2
    for rec in records:
        ws.cell(row=row, column=2, value=rec["Field1"])
        ws.cell(row=row, column=3, value=rec["Field2"])
        ws.cell(row=row, column=4, value=rec["Field3"])
        img_path = rec["Image Path"]
        if os.path.isfile(img_path):
            try:
                pil_img = PILImage.open(img_path)
                orig_width, orig_height = pil_img.size
                fixed_width = 100  # Fixed width for all images
                scale = fixed_width / orig_width
                fixed_height = int(orig_height * scale)
                xl_img = XLImage(img_path)
                xl_img.width = fixed_width
                xl_img.height = fixed_height
                cell_coordinate = f"A{row}"
                ws.add_image(xl_img, cell_coordinate)
            except Exception as e:
                ws.cell(row=row, column=1, value="Error loading image")
        else:
            ws.cell(row=row, column=1, value="Image not found")
        row += 1
    excel_file = "records.xlsx"
    try:
        wb.save(excel_file)
        messagebox.showinfo("Export", f"Excel file exported as {excel_file}")
    except Exception as e:
        messagebox.showerror("Export Error", f"Failed to export Excel: {e}")

# ------------------
# Setup the main window and frames

root = tk.Tk()
root.title("Image & Data Viewer")
root.geometry("600x700")

# Container frame to hold view and input frames
container = tk.Frame(root)
container.pack(fill="both", expand=True)

# --- View Frame ---
view_frame = tk.Frame(container)
view_frame.place(relx=0, rely=0, relwidth=1, relheight=1)

view_image_label = tk.Label(view_frame, text="No record available", bg="white", bd=2, relief=tk.SUNKEN)
view_image_label.pack(pady=10)

view_field1_label = tk.Label(view_frame, text="", font=default_font)
view_field1_label.pack(pady=2)
view_field2_label = tk.Label(view_frame, text="", font=default_font)
view_field2_label.pack(pady=2)
view_field3_label = tk.Label(view_frame, text="", font=default_font)
view_field3_label.pack(pady=2)

nav_frame = tk.Frame(view_frame)
nav_frame.pack(pady=10)

back_button = tk.Button(nav_frame, text="<< Back", command=previous_record, width=10, font=default_font)
back_button.grid(row=0, column=0, padx=5)
forward_button = tk.Button(nav_frame, text="Forward >>", command=next_record, width=10, font=default_font)
forward_button.grid(row=0, column=1, padx=5)
delete_button = tk.Button(nav_frame, text="Delete", command=delete_record, width=10, font=default_font)
delete_button.grid(row=0, column=2, padx=5)
add_new_button = tk.Button(nav_frame, text="Add New", command=switch_to_input_mode, width=10, font=default_font)
add_new_button.grid(row=0, column=3, padx=5)

export_frame = tk.Frame(view_frame)
export_frame.pack(pady=10)
export_pdf_button = tk.Button(export_frame, text="Export PDF", command=export_to_pdf, width=12, font=default_font)
export_pdf_button.grid(row=0, column=0, padx=5)
export_excel_button = tk.Button(export_frame, text="Export Excel", command=export_to_excel, width=12, font=default_font)
export_excel_button.grid(row=0, column=1, padx=5)

# --- Input Frame ---
input_frame = tk.Frame(container)
input_frame.place(relx=0, rely=0, relwidth=1, relheight=1)

select_image_button = tk.Button(input_frame, text="Select Image", command=select_input_image, font=default_font)
select_image_button.pack(pady=10)

input_image_label = tk.Label(input_frame, text="No image selected", bg="white", bd=2, relief=tk.SUNKEN)
input_image_label.pack(pady=10)

tk.Label(input_frame, text="بخش اول :", font=default_font).pack(pady=2)
entry1 = tk.Entry(input_frame, width=50, font=default_font)
entry1.pack(pady=2)

tk.Label(input_frame, text="نام کالا:", font=default_font).pack(pady=2)
entry2 = tk.Entry(input_frame, width=50, font=default_font)
entry2.pack(pady=2)

tk.Label(input_frame, text="تاریخ:", font=default_font).pack(pady=2)
entry3 = tk.Entry(input_frame, width=50, font=default_font)
entry3.pack(pady=2)

input_buttons_frame = tk.Frame(input_frame)
input_buttons_frame.pack(pady=10)

save_button = tk.Button(input_buttons_frame, text="Save", command=save_new_record, width=10, font=default_font)
save_button.grid(row=0, column=0, padx=5)
cancel_button = tk.Button(input_buttons_frame, text="Cancel", command=cancel_input, width=10, font=default_font)
cancel_button.grid(row=0, column=1, padx=5)

# Load records and determine which frame to show initially
load_records()
if records:
    view_frame.tkraise()
    show_record(current_record_index)
else:
    input_frame.tkraise()

root.mainloop()
